import requests
from bs4 import BeautifulSoup
import time
import re
import math
import openpyxl
import concurrent.futures

# URL главной страницы сайта
url = 'https://food.ru'  # Базовый URL сайта

# Категории, ссылки и количество страниц для каждой категории
categories = ['первые блюда', 'вторые блюда', 'закуски', 'салаты', 'гарниры', 'десерты', 'выпечка', 'напитки', 'заготовки и консервы', 'соусы и маринады']
category_links = ['/recipes/pervye-bliuda', '/recipes/vtorye-bliuda', '/recipes/zakuski', '/recipes/salaty', '/recipes/garniry', '/recipes/deserty', '/recipes/vypechka', '/recipes/napitki', '/recipes/zagotovki', '/recipes/sousy-i-marinady']
category_numbers = ['384', '2631', '775', '804', '90', '820', '700', '277', '106', '93']  # Количество страниц

# Функция для итерации по категориям и страницам
def scrape_pages():
	for i in range(len(category_links)):
        # Текущая категория и количество страниц
		category = categories[i]
		link = category_links[i]
		max_page = int(category_numbers[i])  # Преобразуем количество страниц в число

		print(f"Скрапинг категории: {category} (страниц: {max_page})")
        
        # Проходим по каждой странице категории
		for page in range(1, max_page + 1):
            # Формируем ссылку для текущей страницы
			if page == 1:
				page_url = f"{url}{link}"
			else:
				page_url = f"{url}{link}?page={page}"
			print(f"Получение страницы: {page_url}")
            
            # Отправляем GET-запрос на страницу
			response = requests.get(page_url)
			if response.status_code == 200:
				soup = BeautifulSoup(response.text, 'html.parser')

                # Поиск всех <a> с классом 'card_card__YG0I9' на текущей странице
				recipe_links = soup.find_all('a', class_='card_card__YG0I9')
                
                # Проходим по каждой найденной ссылке на рецепт
				for recipe_link in recipe_links:
					recipe_href = recipe_link.get('href')
					if recipe_href:
                        # Формируем полный URL для страницы рецепта
						recipe_page_url = f"{url}{recipe_href}"
		##				#print(f"\n\n\nПолучение страницы рецепта: {recipe_page_url}")
                        
                        # Отправляем GET-запрос на страницу рецепта
						recipe_response = requests.get(recipe_page_url)
						if recipe_response.status_code == 200:
							recipe_soup = BeautifulSoup(recipe_response.text, 'html.parser')

                            # Здесь можно добавить логику обработки страницы рецепта
                            # Например, выводим заголовок страницы рецепта
							recipe_title = recipe_soup.title.string if recipe_soup.title else "Без названия"
		##					#print(f"Заголовок рецепта: {recipe_title}")





							recipe_details = parse_recipe_details(recipe_soup, category)
							#print(recipe_details, end='\n')

							update_excel_with_recipe(recipe_details, 'dishes.xlsx')
						else:
							print(f"Не удалось получить страницу рецепта {recipe_page_url}, статус: {recipe_response.status_code}")
                        
                        # Задержка между запросами, чтобы снизить нагрузку на сервер
						#time.sleep(0.1)
			else:
				print(f"Не удалось получить страницу {page_url}, статус: {response.status_code}")





def clean_recipe_soup(recipe_soup):
    # Преобразуем объект soup в строку
    soup_str = str(recipe_soup)
    
    # Заменяем все множественные пробелы на один пробел
    soup_cleaned = re.sub(r'\s+', ' ', soup_str)
    
    # Возвращаем обратно в объект BeautifulSoup
    return BeautifulSoup(soup_cleaned, 'html.parser')


def parse_recipe_details(recipe_soup, category):
    recipe_soup = BeautifulSoup(str(recipe_soup).replace('<!-- -->', ' '), 'html.parser')
    recipe_soup = BeautifulSoup(str(recipe_soup).replace('\xa0', ' '), 'html.parser')
    
    recipe_soup = clean_recipe_soup(recipe_soup)
    # Словарь для хранения всех данных
    recipe_data = {}
    
    # 1. Извлечение заголовка рецепта (<h1> с классом 'title_main__ok7t1')
    title = recipe_soup.find('h1', class_='title_main__ok7t1')
    recipe_data['title'] = title.text.strip() if title else "Заголовок не найден"
    
    # 2. Готовая переменная category (категория)
    recipe_data['category'] = category
    
    # 3. Извлечение описания (<span> с классом 'markup_text__F9WKe')
    description = recipe_soup.find('span', class_='markup_text__F9WKe')
    recipe_data['description'] = description.text.strip() if description else "Описание не найдено"
    
    # 4. Извлечение словаря питательных веществ (<span> с классом 'nutrient_title__JDSmX' и 'nutrient_value__dd48k')
    nutrient_info = {}
    nutrients = recipe_soup.find_all('span', class_='nutrient_title__JDSmX')
    values = recipe_soup.find_all('span', class_='nutrient_value__dd48k')
    
    if len(nutrients) == len(values):
        for nutrient, value in zip(nutrients, values):
            nutrient_info[nutrient.text.strip()] = value.text.strip()
    recipe_data['nutrients'] = nutrient_info
    
    # 5. Извлечение ингредиентов и их количеств (словари <span class="ingredientsTable_text__3ILFA"> и <span class="value">)
    ingredients_info = {}
    ingredients = recipe_soup.find_all(['a', 'span'], class_='ingredientsTable_text__3ILFA')
    quantities = []

    even_ingredients = []
    for i in range(len(ingredients)):
        if i % 2 != 0:  # Если индекс чётный
            even_ingredients.append(ingredients[i])  # Добавляем в quantities
            quantities.append(ingredients[i])  # Добавляем элемент в quantities

# Удаляем чётные элементы из ingredients
    ingredients = [ingredient for i, ingredient in enumerate(ingredients) if i % 2 == 0]

    if len(ingredients) == len(quantities):
        for ingredient, quantity in zip(ingredients, quantities):
        # Спускаемся в самый глубокий вложенный элемент внутри тега ingredient
            deepest_element = ingredient
            while deepest_element.find(True):
                deepest_element = deepest_element.find(True)
        
            # Извлекаем текст из самого глубокого элемента
            ingredient_text = deepest_element.text.strip()
            # Обработка количества, если содержит знак '=', отсекаем всё слева
            quantity_text = quantity.text.strip()
            if '=' in quantity_text:
                quantity_text = quantity_text.split('=')[-1]
        
            # Убираем все символы, которые не являются цифрами
            if re.search(r'\d', quantity_text):
                quantity_clean = ''.join(re.findall(r'\d+', quantity_text))

    # Проверяем наличие точки
                if '.' in quantity_text:
                    try:
            # Преобразуем в целое число, делим на 10 и обратно в строку
                        quantity_clean = str(int(quantity_clean) // 10)
                    except ValueError:
            # Обработка ошибки преобразования в целое число
                        pass
            else:
                quantity_clean = quantity_text
            # Записываем в словарь ингредиентов
            ingredients_info[ingredient_text] = '1' if quantity_clean == '0' else quantity_clean

# Сохраняем результат в словарь данных рецепта
    recipe_data['ingredients'] = ingredients_info
    
    # 6. Извлечение свойств блюда (<div class="properties_value__kAeD9">)
    properties = recipe_soup.find_all('div', class_='properties_value__kAeD9')
    if properties:
        last_property = properties[-1].text.strip()  # Получаем последний элемент
        recipe_data['properties'] = last_property
    else:
        recipe_data['properties'] = "Свойства не найдены"

    
    # 7. Извлечение значения порции (<input class="input yield">)
    flag = False
    servings_input = recipe_soup.find('input', class_='input yield default yield')

    if servings_input:
        servings_value = servings_input.get('value')
        #print(servings_value)
        
        if servings_value.isdigit():
            servings_value = int(servings_value)
            
            # Коррекция значений ингредиентов (каждый четный элемент словаря делим на servings_value)
            for i, (ingredient, quantity) in enumerate(ingredients_info.items()):
                if 1:  # Только четные элементы (индекс с 1)
                    try:
                        corrected_quantity = math.ceil(int(quantity) / servings_value)
                        ingredients_info[ingredient] = corrected_quantity
                    except (ValueError, ZeroDivisionError):
                        continue
    

    # 8. записываем картинку 
    img_tags = recipe_soup.find_all('img', alt=recipe_data['title'])
    
    if img_tags:
        # Берем последний найденный тег <img>
        last_img_tag = img_tags[-1]
        # Извлекаем значение атрибута src
        recipe_data['image'] = last_img_tag.get('src')
    else:
        recipe_data['image'] = ''
    # Обновляем данные ингредиентов в словаре
    #recipe_data['ingredients_corrected'] = ingredients_info


    #9. считываем рецепт

    section = recipe_soup.find('section', id='step-by-step-recipe')

    if section:
    # Ищем все теги <span> с классом 'markup_text__F9WKe' внутри найденной секции
        spans = section.find_all('span', class_='markup_text__F9WKe')

        if spans and len(spans) > 1:
        # Пропускаем первый элемент и собираем текст из оставшихся
            combined_text = '\n'.join([span.text.strip() for span in spans[1:]])
            recipe_data['recipe'] = combined_text.strip()  # Убираем лишние пробелы
        else:
            recipe_data['recipe'] = "Описание не найдено"
    else:
        recipe_data['recipe'] = "Секция с рецептом не найдена"




    #10. считываем время
    ready_time = recipe_soup.find('meta', itemprop='totalTime')
    cooking_time = recipe_soup.find('meta', itemprop='prepTime')

    # Извлечение значений из атрибутов "content"
    ready_minutes = ready_time['content'] if ready_time else 'PT0M'
    cooking_minutes = cooking_time['content'] if cooking_time else 'PT0M'

    # Преобразование времени из формата PTxxM в минуты
    ready_minutes_value = int(ready_minutes.replace('PT', '').replace('M', '')) if ready_minutes else 0
    cooking_minutes_value = int(cooking_minutes.replace('PT', '').replace('M', '')) if cooking_minutes else 0

    # Сохранение времени в строковом формате
    time_data = f"{{'ready': {ready_minutes_value}, 'cooking_time': {cooking_minutes_value}}}"
    recipe_data['time'] = time_data
    
    return recipe_data

# Пример вызова функции:
# recipe_details = parse_recipe_details(recipe_soup, category)
# print(recipe_details)








def update_excel_with_recipe(recipe_details, file_path):
    """
    Обновляет Excel-файл, добавляя новые данные рецепта, если такого блюда ещё нет в таблице.

    :param recipe_details: словарь с информацией о рецепте (title, category, description, nutries, ingredients, properties)
    :param file_path: путь к Excel-файлу
    """
    
    # Открываем Excel-файл и активный лист
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Получаем данные из столбца 'dish' (столбец А)
    dish_column = [cell.value for cell in sheet['A']]
    
    # Название блюда, которое нужно проверить
    dish_name = recipe_details['title']
    
    # Проверка: если такое блюдо уже есть в таблице, не добавляем его
    if dish_name in dish_column:
        print(f"Блюдо '{dish_name}' уже существует в таблице.")
        return
    
    # Находим первую пустую строку
    empty_row = len(dish_column) + 1
    
    # Заполняем новую строку данными
    sheet.cell(row=empty_row, column=1, value=recipe_details['title'])       # dish
    sheet.cell(row=empty_row, column=2, value=recipe_details['category'])    # category
    sheet.cell(row=empty_row, column=3, value=recipe_details['description']) # description
    sheet.cell(row=empty_row, column=4, value=str(recipe_details['nutrients']))     # macros
    sheet.cell(row=empty_row, column=5, value=str(recipe_details['ingredients'])) # ingredients
    sheet.cell(row=empty_row, column=6, value=recipe_details['properties'])  # allergy
    sheet.cell(row=empty_row, column=7, value=recipe_details['image'])  # image src
    sheet.cell(row=empty_row, column=8, value=recipe_details['recipe'])  # рецепт
    sheet.cell(row=empty_row, column=9, value=recipe_details['time'])  # время

    
    # Сохраняем изменения
    workbook.save(file_path)
    #print(f"Блюдо '{dish_name}' добавлено в таблицу.")










scrape_pages()

'''
def execute_foo_multithreaded():
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = []
        
        # Запускаем функцию foo() в нескольких потоках
        for _ in range(10):  # Запустим 10 потоков
            futures.append(executor.submit(scrape_pages))
        
        # Ждем завершения всех потоков
        for future in concurrent.futures.as_completed(futures):
            try:
                future.result()  # Проверяем на ошибки выполнения
            except Exception as exc:
                print(f"Поток завершился с ошибкой: {exc}")

execute_foo_multithreaded()

'''