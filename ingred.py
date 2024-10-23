import requests
from bs4 import BeautifulSoup
import time
import re
import math
import openpyxl
import concurrent.futures

url = 'https://food.ru/'  # Базовый URL сайта

category_number = 57

def scrape_pages():
        # Проходим по каждой странице категории
		for page in range(48, category_number + 1):
			print(f"Скрапинг категории: продукты, страница: {page}")
            # Формируем ссылку для текущей страницы
			if page == 1:
				page_url = f"{url}/products"
			else:
				page_url = f"{url}/products?page={page}"
			print(f"Получение страницы: {page_url}")
            
            # Отправляем GET-запрос на страницу
			response = requests.get(page_url)
			if response.status_code == 200:
				soup = BeautifulSoup(response.text, 'html.parser')

                # Поиск всех <a> с классом 'card_card__YG0I9' на текущей странице
				recipe_links = soup.find_all('a', class_='productCard_productCard__Z57rS')
                
                # Проходим по каждой найденной ссылке на рецепт
				for recipe_link in recipe_links:
					recipe_href = recipe_link.get('href')
					if recipe_href:
                        # Формируем полный URL для страницы рецепта
						recipe_page_url = f"{url}{recipe_href}"
		##				#print(f"\n\n\nПолучение страницы рецепта: {recipe_page_url}")
                        
                        # Отправляем GET-запрос на страницу рецепта
						recipe_response = requests.get(recipe_page_url)
						if recipe_response.status_code == 200:
							recipe_soup = BeautifulSoup(recipe_response.text, 'html.parser')

                            # Здесь можно добавить логику обработки страницы рецепта
                            # Например, выводим заголовок страницы рецепта
							recipe_title = recipe_soup.title.string if recipe_soup.title else "Без названия"
		##					#print(f"Заголовок рецепта: {recipe_title}")





							recipe_details = parse_recipe_details(recipe_soup)
							#print(recipe_details, end='\n')

							update_excel_with_recipe(recipe_details, 'ingredients.xlsx')
						else:
							print(f"Не удалось получить страницу рецепта {recipe_page_url}, статус: {recipe_response.status_code}")
                        
                        # Задержка между запросами, чтобы снизить нагрузку на сервер
						#time.sleep(0.1)
			else:
				print(f"Не удалось получить страницу {page_url}, статус: {response.status_code}")



def clean_recipe_soup(recipe_soup):
    # Преобразуем объект soup в строку
    soup_str = str(recipe_soup)
    
    # Заменяем все множественные пробелы на один пробел
    soup_cleaned = re.sub(r'\s+', ' ', soup_str)
    
    # Возвращаем обратно в объект BeautifulSoup
    return BeautifulSoup(soup_cleaned, 'html.parser')


def parse_recipe_details(recipe_soup):
    recipe_soup = BeautifulSoup(str(recipe_soup).replace('<!-- -->', ' '), 'html.parser')
    recipe_soup = BeautifulSoup(str(recipe_soup).replace('\xa0', ' '), 'html.parser')
    
    recipe_soup = clean_recipe_soup(recipe_soup)
    # Словарь для хранения всех данных
    recipe_data = {}
    
    # 1. Извлечение заголовка рецепта (<h1> с классом 'title_main__ok7t1')
    title = recipe_soup.find('h1', class_='title_title__OHmf9')
    recipe_data['title'] = title.text.strip() if title else "Заголовок не найден"

    
    # 3. Извлечение описания (<span> с классом 'markup_text__F9WKe')
    description_wrapper = recipe_soup.find('div', class_='markup_wrapper__1wSxB')
    
    if description_wrapper:
        # Собираем весь текст из вложенных тегов
        description_text = ' '.join([span.text.strip() for span in description_wrapper.find_all('span')])
        recipe_data['description'] = description_text.strip()  # Удаляем лишние пробелы в начале и конце
    else:
        recipe_data['description'] = ""
    
    # 4. Извлечение словаря питательных веществ (<span> с классом 'nutrient_title__JDSmX' и 'nutrient_value__dd48k')
    nutrient_info = {}
    nutrients = recipe_soup.find_all('span', class_='nutrient_title__JDSmX')
    values = recipe_soup.find_all('span', class_='nutrient_value__dd48k')
    
    if len(nutrients) == len(values):
        for nutrient, value in zip(nutrients, values):
            nutrient_info[nutrient.text.strip()] = value.text.strip()
    recipe_data['macros'] = nutrient_info

 
    
    return recipe_data

# Пример вызова функции:
# recipe_details = parse_recipe_details(recipe_soup, category)
# print(recipe_details)


def update_excel_with_recipe(recipe_details, file_path):
    """
    Обновляет Excel-файл, добавляя новые данные рецепта, если такого блюда ещё нет в таблице.

    :param recipe_details: словарь с информацией о рецепте (title, category, description, nutries, ingredients, properties)
    :param file_path: путь к Excel-файлу
    """
    
    # Открываем Excel-файл и активный лист
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Получаем данные из столбца 'dish' (столбец А)
    dish_column = [cell.value for cell in sheet['A']]
    
    # Название блюда, которое нужно проверить
    dish_name = recipe_details['title']
    
    # Проверка: если такое блюдо уже есть в таблице, не добавляем его
    if dish_name in dish_column:
        print(f"Блюдо '{dish_name}' уже существует в таблице.")
        return
    
    # Находим первую пустую строку
    empty_row = len(dish_column) + 1
    
    # Заполняем новую строку данными
    sheet.cell(row=empty_row, column=1, value=recipe_details['title'])       # dish
    sheet.cell(row=empty_row, column=2, value=recipe_details['description']) # description
    sheet.cell(row=empty_row, column=3, value=str(recipe_details['macros']))     # macros
    # Сохраняем изменения
    workbook.save(file_path)




scrape_pages()